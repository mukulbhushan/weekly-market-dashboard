import openpyxl

wb = openpyxl.load_workbook('WEEKLY REPORT SPREADSHEET.xlsx', data_only=True)
with open('excel_structure.txt', 'w', encoding='utf-8') as f:
    for name in wb.sheetnames:
        s = wb[name]
        f.write(f"================ SHEET: {name} (Rows: {s.max_row}, Cols: {s.max_column}) ================\n")
        for r in range(1, min(25, s.max_row+1)):
            row_vals = [s.cell(r, c).value for c in range(1, min(12, s.max_column+1))]
            if any(v is not None for v in row_vals):
                f.write(f"Row {r:2d}: {row_vals}\n")
        f.write("\n")

print("Saved to excel_structure.txt")
