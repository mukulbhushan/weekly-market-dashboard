import streamlit as st
import asyncio
import os
import glob
import datetime
import pandas as pd
import openpyxl
import streamlit.components.v1 as components
from playwright.async_api import async_playwright
import update_spreadsheet

st.set_page_config(
    page_title="Executive Weekly Market Dashboard",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for Executive Theme
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Newsreader:wght@600&family=Inter:wght@400;500;600;700&family=IBM+Plex+Mono:wght@500;600;700&display=swap');
    
    .stApp {
        background-color: #F8FAFC;
        font-family: 'Inter', sans-serif;
    }
    .mono { font-family: 'IBM Plex Mono', monospace; }
    
    .spot-strip {
        background: #FFFFFF;
        border: 1px solid #CBD5E1;
        border-left: 5px solid #B45309;
        border-radius: 6px;
        padding: 10px 16px;
        margin-bottom: 16px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    .spot-badge {
        font-family: 'IBM Plex Mono', monospace;
        font-size: 11px;
        font-weight: 700;
        color: #B45309;
        text-transform: uppercase;
        letter-spacing: 0.08em;
    }
    .metric-card {
        background: #FFFFFF;
        border: 1px solid #E2E8F0;
        border-radius: 6px;
        padding: 12px 16px;
        box-shadow: 0 1px 2px rgba(0,0,0,0.03);
    }
    div[data-testid="stMetricValue"] {
        font-family: 'IBM Plex Mono', monospace;
        font-weight: 700;
        font-size: 22px;
    }
    div[data-testid="stMetricDelta"] {
        font-family: 'IBM Plex Mono', monospace;
        font-weight: 600;
        font-size: 13px;
    }
</style>
""", unsafe_allow_html=True)

# Helper to generate page preview images if needed
async def ensure_preview_images():
    if os.path.exists("weekly-market-dashboard.html"):
        try:
            async with async_playwright() as p:
                browser = None
                for channel in ['msedge', 'chrome', None]:
                    try:
                        if channel: browser = await p.chromium.launch(channel=channel, headless=True)
                        else: browser = await p.chromium.launch(headless=True)
                        break
                    except Exception: continue
                if browser:
                    page = await browser.new_page(viewport={'width': 1920, 'height': 1200})
                    html_url = 'file:///' + os.path.abspath('weekly-market-dashboard.html').replace('\\', '/')
                    await page.goto(html_url, wait_until='networkidle')
                    await page.wait_for_timeout(800)
                    await page.locator('#page1Section').screenshot(path='page1_preview.png')
                    await page.locator('#page2Section').screenshot(path='page2_preview.png')
                    await browser.close()
        except Exception as err:
            print("Preview capture notice:", err)

# Sidebar Actions & Controls
with st.sidebar:
    if os.path.exists("raru_logo.png"):
        st.image("raru_logo.png", width=180)
    st.title("Market Dashboard")
    st.caption("Executive Research & Portfolio Desk")
    st.divider()
    
    st.subheader("⚡ Live Operations")
    
    if st.button("🔄 Refresh Live Market Data", use_container_width=True, type="primary"):
        with st.spinner("Fetching live market feeds, calculating MTD flows & updating spreadsheet..."):
            try:
                asyncio.run(update_spreadsheet.main_pipeline())
                asyncio.run(ensure_preview_images())
                st.success("✅ Dashboard & Excel synchronized successfully!")
                st.rerun()
            except Exception as e:
                st.error(f"Execution Error: {e}")

    # Locate latest PDF
    pdf_files = sorted(glob.glob("Weekly_Market_Dashboard_*.pdf"), reverse=True)
    latest_pdf = pdf_files[0] if pdf_files else None
    
    if latest_pdf and os.path.exists(latest_pdf):
        with open(latest_pdf, "rb") as f:
            pdf_bytes = f.read()
        st.download_button(
            label="📥 Download Executive PDF (2-Page)",
            data=pdf_bytes,
            file_name=os.path.basename(latest_pdf),
            mime="application/pdf",
            use_container_width=True
        )
        st.caption(f"Ready: `{os.path.basename(latest_pdf)}` ({len(pdf_bytes):,} bytes)")

    if os.path.exists("WEEKLY REPORT SPREADSHEET.xlsx"):
        with open("WEEKLY REPORT SPREADSHEET.xlsx", "rb") as f:
            excel_bytes = f.read()
        st.download_button(
            label="📊 Download Excel Spreadsheet",
            data=excel_bytes,
            file_name="WEEKLY REPORT SPREADSHEET.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# Main Title Header
col_hdr1, col_hdr2 = st.columns([3, 1])
with col_hdr1:
    st.title("Executive Weekly Market Dashboard")
    st.caption("Live Benchmarks · Institutional Capital Flows · Macro Read & Valuation Metrics")
with col_hdr2:
    st.markdown(f"""
    <div style="text-align:right; font-family:'IBM Plex Mono', monospace; font-size:12px; color:#475569; padding-top:10px;">
        <span style="background:#FEF3C7; color:#B45309; border:1px solid #B45309; padding:3px 8px; border-radius:3px; font-weight:700;">
            ACTIVE BRIEFING
        </span><br>
        Updated: {datetime.datetime.now().strftime('%d %b %Y, %I:%M %p')}
    </div>
    """, unsafe_allow_html=True)

st.write("")

# Top Metric Strip (Nifty 50, Sensex, Bank Nifty, Fin Nifty)
try:
    wb = openpyxl.load_workbook("WEEKLY REPORT SPREADSHEET.xlsx", data_only=True)
    idx_sheet = wb["Executive Dashboard & Index Ben"]
    
    n50_close = idx_sheet.cell(2, 2).value or 24330.30
    n50_chg = (idx_sheet.cell(2, 4).value or -0.0015) * 100
    n50_pts = idx_sheet.cell(2, 3).value or -35.70
    
    bnf_close = idx_sheet.cell(3, 2).value or 57571.50
    bnf_chg = (idx_sheet.cell(3, 4).value or 0.0014) * 100
    bnf_pts = idx_sheet.cell(3, 3).value or 80.40
    
    fin_close = idx_sheet.cell(4, 2).value or 26253.95
    fin_chg = (idx_sheet.cell(4, 4).value or 0.0015) * 100
    fin_pts = idx_sheet.cell(4, 3).value or 40.30
    
    snx_close = idx_sheet.cell(5, 2).value or 77797.64
    snx_chg = (idx_sheet.cell(5, 4).value or -0.0027) * 100
    snx_pts = idx_sheet.cell(5, 3).value or -211.61
except Exception:
    n50_close, n50_pts, n50_chg = 24330.30, -35.70, -0.15
    snx_close, snx_pts, snx_chg = 77797.64, -211.61, -0.27
    bnf_close, bnf_pts, bnf_chg = 57571.50, 80.40, 0.14
    fin_close, fin_pts, fin_chg = 26253.95, 40.30, 0.15

m1, m2, m3, m4 = st.columns(4)
m1.metric("NIFTY 50 SPOT", f"{n50_close:,.2f}", f"{n50_pts:+,.2f} ({n50_chg:+.2f}%)")
m2.metric("SENSEX SPOT", f"{snx_close:,.2f}", f"{snx_pts:+,.2f} ({snx_chg:+.2f}%)")
m3.metric("BANK NIFTY SPOT", f"{bnf_close:,.2f}", f"{bnf_pts:+,.2f} ({bnf_chg:+.2f}%)")
m4.metric("FIN NIFTY SPOT", f"{fin_close:,.2f}", f"{fin_pts:+,.2f} ({fin_chg:+.2f}%)")

st.write("")

# Navigation Tabs
tab1, tab2, tab3 = st.tabs([
    "🖥️ Interactive Executive Dashboard", 
    "📄 PDF Report Pages & Export",
    "📊 Excel Data & Institutional Explorer"
])

with tab1:
    st.subheader("Interactive Market Briefing")
    st.caption("Live interactive dashboard with interactive charts, sector rotation heatmap, and institutional flows.")
    if os.path.exists("weekly-market-dashboard.html"):
        with open("weekly-market-dashboard.html", "r", encoding="utf-8") as f:
            html_content = f.read()
        components.html(html_content, height=1400, scrolling=True)
    else:
        st.info("HTML Dashboard not found. Click 'Refresh Live Market Data' in the sidebar to generate it.")

with tab2:
    st.subheader("Investor 2-Page Print-Ready PDF")
    st.caption("Inspected and verified 2-page landscape format designed for weekly investor dissemination.")
    
    col_pdf1, col_pdf2 = st.columns([1, 1])
    
    if latest_pdf and os.path.exists(latest_pdf):
        with open(latest_pdf, "rb") as f:
            pdf_bytes = f.read()
        
        st.download_button(
            label="⬇️ Download Executive 2-Page PDF Document",
            data=pdf_bytes,
            file_name=os.path.basename(latest_pdf),
            mime="application/pdf",
            type="primary",
            use_container_width=True
        )
    
    st.write("")
    
    # Render preview images of Page 1 and Page 2
    if os.path.exists("page1_preview.png") and os.path.exists("page2_preview.png"):
        p_col1, p_col2 = st.columns(2)
        with p_col1:
            st.markdown("#### Page 1: Benchmark Summary & Sector Performance")
            st.image("page1_preview.png", use_container_width=True)
        with p_col2:
            st.markdown("#### Page 2: Institutional Flows, Bulk Deals & Macro Events")
            st.image("page2_preview.png", use_container_width=True)

with tab3:
    st.subheader("Underlying Excel Sheets Explorer")
    if os.path.exists("WEEKLY REPORT SPREADSHEET.xlsx"):
        wb = openpyxl.load_workbook("WEEKLY REPORT SPREADSHEET.xlsx", data_only=True)
        sheet_choice = st.selectbox("Select Sheet to View:", wb.sheetnames)
        if sheet_choice:
            ws = wb[sheet_choice]
            data = list(ws.iter_rows(values_only=True))
            if data and len(data) > 0:
                header = [str(c) if c is not None else f"Col {i+1}" for i, c in enumerate(data[0])]
                rows = data[1:]
                df = pd.DataFrame(rows, columns=header)
                st.dataframe(df, use_container_width=True)
    else:
        st.info("Spreadsheet not found. Please refresh live market data first.")
